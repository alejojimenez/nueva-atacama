from codigo.app_atacama import Scraper_Atacama
#import smtplib
from openpyxl import load_workbook

def send_notification():
    # Código para enviar correo electrónico de notificación
    print('')

if __name__ == '__main__':
    
    print('Obteniendo credenciales...')
    print('----------------------------------------------------------------------')
        
    credencials = '.\\config\\credenciales.xlsx'
    libro_accesos = load_workbook(credencials)
    hoja_credenciales = libro_accesos['Hoja1']
        
    for j in hoja_credenciales.iter_rows(2):
        try:
            rut = j[0].value
            passw = j[1].value
            web = j[2].value
            break
        except:
            ('no hay credenciales')
            
    email = rut
    password = passw
    url = web
    driver_path = 'chromedriver.exe'
    
    scraper = Scraper_Atacama(url, email, password, driver_path)
    print('ingresamos en la clase Scraper_Atacama...')
    print('----------------------------------------------------------------------')
    
    scraper.login()
    print('Hacemos loginen en el portal...')
    print('----------------------------------------------------------------------')

    scraper.scrapping_codiner()
    print('Hacemos scrapping al portal...')
    print('----------------------------------------------------------------------')
        
    scraper.archivos()
    print('Extraemos datos')
    print('----------------------------------------------------------------------')

    scraper.close()
    print('Cerramos el bot...')
    print('----------------------------------------------------------------------')
