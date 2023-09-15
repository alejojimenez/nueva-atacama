#Librerias de sistema
import os
import time
import shutil
import pandas as pd

#Librerias datos
import fitz
import glob
from openpyxl import load_workbook

#Libreria web scrapping
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.action_chains import ActionChains

class Scraper_Atacama():

    def __init__(self,url, email, password, driver_path):
        print(url, email, password, driver_path)
        self.url = url
        self.email = email
        self.password = password
        self.driver_path = driver_path

    def wait(self, seconds):
        return WebDriverWait(self.driver, seconds)

    def close(self):
        self.driver.close()
        self.driver = None

    def quit(self):
        self.driver.quit()
        self.driver = None    

    def login(self):
        
        driver_exe = '.domain\\chromedriver.exe'
        credencials = '.\\config\\credenciales.xlsx'

        print('Entrando en la funcion login...')
        print('----------------------------------------------------------------------')
        
        #Seteo variables
        email = self.email
        url = self.url
        driver_path = self.driver_path
        password  = self.password
        
        options = webdriver.ChromeOptions()
        options.add_extension('C:\\roda\\nueva-atacama\\config\\33.9_0.crx')
        options.add_experimental_option('prefs', {
        "download.default_directory": "C:\\roda\\nueva-atacama\\input\\", #Change default directory for downloads
        "download.prompt_for_download": False, #To auto download the file
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome
        })
            
        self.driver = webdriver.Chrome(driver_path, options=options)
        self.driver.get(url)
        self.driver.maximize_window()
        self.driver.implicitly_wait(40)

        # Controlar evento alert() Notificacion
        try:
            alert = WebDriverWait(self.driver, 35).until(EC.alert_is_present())
            alert = Alert(self.driver)
            alert.accept()
                    
        except:
            print("No se encontró ninguna alerta.")  

        # Seleccionar usuario dentro de la pagina
        intentos = 0
        usuario = True
        while (usuario):
            try:
                print('Try en la pestaña usuario..', intentos)
                print('----------------------------------------------------------------------')
                intentos += 1
                element_usuario= self.driver.find_element(By.ID, 'username')
                element_usuario.clear()
                element_usuario.click()
                element_usuario.send_keys(email)
                usuario = False
            except:    
                print('Exception en la pestaña usuario')
                print('----------------------------------------------------------------------')
                usuario = intentos <= 3                

        # Seleccionar clave dentro de la pagina
        intentos = 0
        clave = True
        while (clave):
            try:
                print('Try en la pestaña clave..', intentos)
                print('----------------------------------------------------------------------')
                intentos += 1
                time.sleep(5)
                element_clave= self.driver.find_element(By.ID, 'password')
                element_clave.clear()
                element_clave.click()
                element_clave.send_keys(password)
                clave = False   
            except:
                print('Exception en la pestaña clave')
                print('----------------------------------------------------------------------')
                clave = intentos <= 3 

        # Seleccionar clave dentro de la pagina
        intentos = 0
        ingresar = True
        while (ingresar):
            try:
                print('Try en el borton de ingreso...', intentos)
                print('----------------------------------------------------------------------')
                intentos += 1
                time.sleep(5)
                element_ingreso= self.driver.find_element(By.ID, 'submitIngresar')
                if element_ingreso.is_displayed():
                    action_chains = ActionChains(self.driver)
                    action_chains.click(element_ingreso).perform()
                
                #element_clave.click()
                ingresar = False   
            except:
                print('Exception en el borton de ingreso')
                print('----------------------------------------------------------------------')
                ingresar = intentos <= 3

        print('Hicimos login, comenzamos scrapping')
         
    def scrapping_codiner(self):

        print('Entrando en la funcion Scrapping...')
        print('----------------------------------------------------------------------')
        
        folder_path_config = './config/'
        
        # Especifica la ruta de tu archivo Excel
        excel_file = folder_path_config + "clientes.xlsx"

        # Especifica el nombre de la hoja en la que se encuentran los datos
        hoja_excel = "Hoja1"

        # Carga los datos de Excel en un DataFrame
        df = pd.read_excel(excel_file, sheet_name=hoja_excel)
        print('Dataframe ', df)
        print('----------------------------------------------------------------------')
        
        #Esperamos detectar iframe para poder obtener los datos de la tabla
        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.TAG_NAME, "tbody")))
        
        tabla = element_ingreso= self.driver.find_element(By.TAG_NAME, 'tbody')
        filas = tabla.find_elements(By.TAG_NAME, 'tr')
        
        for index,boton in enumerate(filas):
            print(index,'-',boton.get_attribute('innerText'))
        
        largo = len(filas)
        
        #Iteramos para obtener las url de descarga y la fecha
        i=2
        while i < largo+1:

            #Obtenemos la fecha
            intentos = 0
            get_fecha = True
            while (get_fecha):
                try:
                    print('Try en obtener la fecha...', intentos)
                    print('----------------------------------------------------------------------')
                    intentos += 1
                    element_fecha = self.driver.find_element(By.XPATH, f'/html/body/div[7]/main/div/div[1]/div/div[2]/table/tbody/tr[{i}]/td[1]')
                    fecha  = element_fecha.text
                    mes,año = fecha.split()
                    time.sleep(5)

                    get_fecha = False   
                except:
                    print('Exception en el borton de ingreso')
                    print('----------------------------------------------------------------------')
                    get_fecha = intentos <= 3
                    
            #Obtenemos el boton de descarga
            intentos = 0
            get_descarga = True
            while (get_descarga):
                try:
                    print('Try en hacer click sobre boton de descarga...', intentos)
                    print('----------------------------------------------------------------------')
                    intentos += 1
                    element_descarga = self.driver.find_element(By.XPATH, f'/html/body/div[7]/main/div/div[1]/div/div[2]/table/tbody/tr[{i}]/td[2]/a')
                    element_descarga.click()
                    time.sleep(5)
                    get_descarga = False   
                except:
                    print('Exception en hacer click sobre boton de descarga')
                    print('----------------------------------------------------------------------')
                    get_descarga = intentos <= 3
 
            #Entreamos un tiempo para que descargue
            time.sleep(10)

            folder_path = './input/'

            # Cruce datos faltantes para ontener
            for index, row in df.iterrows():
                    
                df_nro_cliente = df.loc[index, 'nro_cliente']
                df_sucursal = df.loc[index, 'sucursal']
                print('Nro. Cliente: ', df_nro_cliente, 'Sucursal: ', df_sucursal)
                print('--------------------------------------------------------------------------')

            #Buscamos el archivo que se haya descargado que comience con el nombre boleta para poder moverlo a la carpeta input
            for filename in os.listdir(folder_path):
                if filename.startswith("RE") and filename.endswith('.pdf'):
                    nombre_archivo = filename
                    print(nombre_archivo)
                    print('Numero factura: ', nombre_archivo[21:-22])
                    print('----------------------------------------------------------------------')
                    try:
                        os.rename(folder_path+nombre_archivo,folder_path+f'{df_nro_cliente}_{nombre_archivo[21:-22]}_{año}.pdf')
                    except:
                        print('archivo corrupto o no disponible')

            #Buscamos el archivo que se haya descargado que comience con el nombre boleta para poder moverlo a la carpeta input
            for filename in os.listdir(folder_path):
                if filename.startswith("RE") and filename.endswith('.pdf'):
                    nombre_archivo = filename
                    print(nombre_archivo)
                    try:
                        os.rename(folder_path+nombre_archivo,folder_path+f'Copiapo_Sur_{mes}_{año}.pdf')
                    except:
                        print('archivo corrupto o no disponible')

            i+=1        
            print('pasamos al siguiente archivo')
            
        print('Terminamos descargas')

    def archivos(self):
        
        folder_path = './input/'
        output_path = './output/'
        
        #Revisamos si hay archivos pdf en la carpeta input
        archivos_pdf = glob.glob(os.path.join(folder_path, '*.pdf'))

        #Si no encuentra archivos es porque no se realizo la ejecucion correcta y hay que mandar mail
        if not archivos_pdf:
            print(f'No se encontraron archivos PDF en la carpeta "{folder_path}".')
        else:
            #Si encuentra me entregara todos los documentos con los que trabajaremos
            print(f'Se encontraron los siguientes archivos PDF en la carpeta "{folder_path}":')
            
            for archivo in archivos_pdf:
                            
                with fitz.open(archivo) as pdf_documento:
                    texto_completo = ''

                    for pagina_num in range(pdf_documento.page_count):
                        pagina = pdf_documento.load_page(pagina_num)
                        texto_completo += pagina.get_text()
                    
                    lista_limpia = [elemento.strip() for elemento in texto_completo.split('\n')]
                    
                    #PRIMERA TABLA
                    #Posicion 0
                    factura_elec_bruto = lista_limpia[2]
                    partes = factura_elec_bruto.split("N°")
                    factura_elec = partes[1].strip()
                    
                    #Posicion 1
                    giro_bruto = lista_limpia[15]
                    partes = giro_bruto.split(":")
                    giro = partes[1].strip()

                    #Posicion 2
                    direccion_bruto = lista_limpia[16]
                    partes = direccion_bruto.split(":")
                    direccion = partes[1].strip()
                    
                    numero_direccion = lista_limpia[17]
                    
                    direccion_completa = direccion + ' ' + numero_direccion
                    
                    #Posicion 3
                    ciudad = lista_limpia[18]
                    #Posicion 4
                    ruta_bruto = lista_limpia[19]
                    partes = ruta_bruto.split(":")
                    ruta = partes[1].strip()

                    #Posicion 5
                    rut_bruto= lista_limpia[20]
                    partes = rut_bruto.split(":")
                    rut = partes[1].strip()
                    
                    #SEGUNDA TABLA
                    #Posicion 6
                    consumo_bruto = lista_limpia[23]
                    partes = consumo_bruto.split("(")
                    consumo = partes[0].strip()
                    
                    #Posicion 7
                    lectura_bruto = lista_limpia[24]
                    lectura_actual = lectura_bruto.split('(')[1].split(')')[0]
                    
                    #Posicion 8
                    valor_lectura_1 = lista_limpia[33]
                    #Posicion 9
                    lectura_bruto2 = lista_limpia[25]
                    lectura_anterior = lectura_bruto2.split('(')[1].split(')')[0]
                    #Posicion 10
                    valor_lectura_2 = lista_limpia[34]
                    #Posicion 11
                    consumo_periodo =lista_limpia[35]
                    #Posicion 12
                    consumo_facturado = lista_limpia[36]
                    
                    #Posicion 13
                    sobreconsumo = lista_limpia[37]
                    if sobreconsumo == '':
                        sobreconsumo = 0
                    elif sobreconsumo != '':
                        sobreconsumo = lista_limpia[37]
                        
                    #Posicion 14
                    agua_punta = lista_limpia[38]
                    if agua_punta == '':
                        agua_punta = 0
                    elif agua_punta != '':
                        agua_punta = lista_limpia[37]
                        
                    valor_a_definir = lista_limpia[42]
                    if valor_a_definir == 'Límite de Sobreconsumo':
                        #Posicion 15
                        limite = lista_limpia[43]
                        #Posicion 16
                        fecha_estimada = lista_limpia[45]
                    elif valor_a_definir == 'Fecha Estimada Próxima Lectura':
                        #Posicion 15
                        limite = 0
                        #Posicion 16
                        fecha_estimada = lista_limpia[43]
                    
                    #TERCERA TABLA
                    #Posicion 17
                    n_servicio = lista_limpia[10]
                    #Posicion 18
                    vencimiento = lista_limpia[12]
                    #Posicion 19
                    total_a_pagar = lista_limpia[14]
                    
                    #CUARTA TABLA
                    #Posicion 20
                    elemento_a_buscar = 'Cargo Fijo'
                    try:
                        posicion_1 = lista_limpia.index(elemento_a_buscar)
                        cargo_fijo_bruto = lista_limpia[posicion_1+1]
                        cargo_fijo = cargo_fijo_bruto.replace("$", "")
                    except:
                        print('elemento no se encuentra disponible')
                        cargo_fijo = 0

                    #Posicion 21
                    elemento_a_buscar = 'Consumo Agua Punta'
                    try:
                        posicion_2 = lista_limpia.index(elemento_a_buscar)
                        consumo_agua_punta_bruto = lista_limpia[posicion_2+1]
                        consumo_agua_punta = consumo_agua_punta_bruto.replace("$", "").replace(".", "").strip()
                    except:
                        print('elemento no se encuentra disponible')
                        consumo_agua_punta = 0

                    #Posicion 22
                    elemento_a_buscar = 'Sobreconsumo agua potable'
                    try:
                        posicion_3 = lista_limpia.index(elemento_a_buscar)
                        sobreconsumo_agua_potable_b = lista_limpia[posicion_3+1]
                        sobreconsumo_agua_potable = sobreconsumo_agua_potable_b.replace("$", "").replace(".", "").strip()
                    except:
                        print('elemento no se encuentra disponible')
                        sobreconsumo_agua_potable = 0

                    #Posicion 23
                    elemento_a_buscar = 'Total Servicios Mes'
                    try:
                        posicion_4 = lista_limpia.index(elemento_a_buscar)
                        total_servicio_mes = lista_limpia[posicion_4+1]
                    except:
                        print('elemento no se encuentra disponible')
                        total_servicio_mes = '$0'
        
                    #Posicion 24
                    elemento_a_buscar = 'Amortizacion Servicio'
                    try:
                        posicion_5 = lista_limpia.index(elemento_a_buscar)
                        amortizacion_bruto = lista_limpia[posicion_5+1]
                        amortizacion = amortizacion_bruto.replace('$','').replace('.','')
                    except:
                        print('elemento no se encuentra disponible')
                        amortizacion =0
                    
                    #Posicion 25
                    elemento_a_buscar = 'Total Venta Mes'
                    try:
                        posicion_6 = lista_limpia.index(elemento_a_buscar)
                        total_venta_mes = lista_limpia[posicion_6+1]
                    except:
                        print('elemento no se encuentra disponible')
                        total_venta_mes = '$0'

                    #Posicion 26
                    elemento_a_buscar = 'Valor Neto'
                    try:
                        posicion_7 = lista_limpia.index(elemento_a_buscar)
                        valor_neto_b = lista_limpia[posicion_7+1]
                        valor_neto = valor_neto_b.replace("$", "").replace('.','').strip()
                        
                    except:
                        print('elemento no se encuentra disponible')
                        valor_neto = 0
                    
                    #Posicion 27
                    elemento_a_buscar = 'IVA 19%'
                    try:
                        posicion_8 = lista_limpia.index(elemento_a_buscar)
                        iva_b = lista_limpia[posicion_8+1]
                        iva = iva_b.replace("$", "").replace('.','').strip()
                    except:
                        print('elemento no se encuentra disponible')
                        iva = 0
                        
                    #Posicion 28
                    elemento_a_buscar = 'Total Docto'
                    try:
                        posicion_9 = lista_limpia.index(elemento_a_buscar)
                        total_docto = lista_limpia[posicion_9+1]
                    except:
                        print('elemento no se encuentra disponible')
                        total_docto = '$0'

                    #Posicion 29
                    elemento_a_buscar = 'TOTAL A PAGAR'
                    try:
                        posicion_10 = lista_limpia.index(elemento_a_buscar)
                        total_a_pagar_final_b = lista_limpia[posicion_10+1]
                        total_a_pagar_final = total_a_pagar_final_b.replace("$", "").replace('.','').strip()
                    except:
                        print('elemento no se encuentra disponible')
                        total_a_pagar_final = '$0'                     

                    #QUINTA TABLA
                    #Posicion 30
                    texto_a_verificar = 'Tarifa Publicada Diario'
                    posicion_11 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_11 = idx
                            tarifa_publicada_bruto = lista_limpia[posicion_11]
                            partes = tarifa_publicada_bruto.split()
                            indice_fecha = next((i for i, parte in enumerate(partes) if '/' in parte), None)
                            if indice_fecha is not None:
                                tarifa_publicada_b = partes[indice_fecha]
                                tarifa_publicada = tarifa_publicada_b.replace('/','-')
                                break
                        
                    #Posicion 31
                    texto_a_verificar = 'Tarifas Incluyen'
                    posicion_12 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_12 = idx
                            grupo_tarifario_bruto = lista_limpia[posicion_12]
                            partes = grupo_tarifario_bruto.split("Grupo Tarifario")
                            grupo_tarifario = partes[-1].strip()
                            break
                        
                    #Posicion 32
                    texto_a_verificar = 'Cargo fijo: $'
                    posicion_13 = None  

                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_13 = idx
                            cargo_fijo_bruto = lista_limpia[posicion_13]
                            partes = cargo_fijo_bruto.split(':')
                            cargo_fijo_final = partes[-1].strip()
                            break
                    
                    
                    #Posicion 33
                    texto_a_verificar = 'Metro cúbico (m3) Agua Potable'
                    posicion_14 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_14 = idx
                            m3_agua_potable_bruto = lista_limpia[posicion_14]
                            partes = m3_agua_potable_bruto.split(':')
                            m3_agua_potable_bruto = partes[-1].strip()
                            m3_agua_potable = m3_agua_potable_bruto.replace("$", "").strip()
                            
                            break
                    
                    #Posicion 34
                    texto_a_verificar = 'Metro cúbico (m3) Agua Potable periodo punta:'
                    posicion_15 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_15 = idx
                            m3_agua_potable_punta_bruto = lista_limpia[posicion_15]
                            partes = m3_agua_potable_punta_bruto.split(':')
                            m3_agua_potable_punta_bruto = partes[-1].strip()
                            m3_agua_potable_punta = m3_agua_potable_punta_bruto.replace("$", "").strip()
                            break
                    
                    #Posicion 35
                    texto_a_verificar = 'Metro cúbico (m3) Agua Potable sobreconsumo:'
                    posicion_16 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_16 = idx
                            m3_agua_potable_sobreconsumo_bruto = lista_limpia[posicion_16]
                            partes = m3_agua_potable_sobreconsumo_bruto.split(':')
                            m3_agua_potable_sobreconsumo_b = partes[-1].strip()
                            m3_agua_potable_sobreconsumo = m3_agua_potable_sobreconsumo_b.replace("$", "").strip()
                            break
                    
                    #Posicion 36
                    texto_a_verificar = 'Metro cúbico (m3) alcantarillado:'
                    posicion_17 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_17 = idx
                            m3_alcantarillado_bruto = lista_limpia[posicion_17]
                            partes = m3_alcantarillado_bruto.split(':')
                            m3_alcantarillado_b = partes[-1].strip()
                            m3_alcantarillado = m3_alcantarillado_b.replace("$", "").strip()
                            break
                    
                    #Posicion 37
                    texto_a_verificar = 'Metro cúbico (m3) tratamiento aguas servidas:'
                    posicion_18 = None 
                    
                    for idx, elemento in enumerate(lista_limpia):
                        if texto_a_verificar in elemento:
                            posicion_18 = idx
                            m3_tratamiento_bruto = lista_limpia[posicion_18]
                            partes = m3_tratamiento_bruto.split(':')
                            m3_tratamiento_b = partes[-1].strip()
                            m3_tratamiento = m3_tratamiento_b.replace("$", "").strip()
                            break
                    
                    #Posicion 38      
                    elemento_a_buscar = 'Factor de Cobro'
                    try:
                        posicion_19 = lista_limpia.index(elemento_a_buscar)
                        factor_cobro = lista_limpia[posicion_19+1] 
                    except:
                        print('elemento no se encuentra disponible')
                        factor_cobro = 0
                        
                    #Posicion 39
                    elemento_a_buscar = 'N° de Medidor'
                    try:
                        posicion_20 = lista_limpia.index(elemento_a_buscar)
                        n_medidor = lista_limpia[posicion_20+1] 
                    except:
                        print('elemento no se encuentra disponible')
                        n_medidor = 0
                    
                    #Posicion 40
                    elemento_a_buscar = 'Diametro de Medidor'
                    try:
                        posicion_21 = lista_limpia.index(elemento_a_buscar)
                        diametro_medidor = lista_limpia[posicion_21+1] 
                    except:
                        print('elemento no se encuentra disponible')
                        diametro_medidor = '$0' 

                #Cargamos libro excel donde volcaremos los datos
                libro = load_workbook(output_path+'/'+'Formato Planilla.xlsx')
                hoja_agua = libro['Agua']
                    
                ultima_fila = hoja_agua.max_row
                
                #Los datos mas importantes
                hoja_agua.cell(row=ultima_fila+1,column=1).value = 1
                hoja_agua.cell(row=ultima_fila+1,column=2).value = n_servicio
                
                #Primera tabla traspasada a excel
                hoja_agua.cell(row=ultima_fila+1,column=3).value = int(factura_elec)
                hoja_agua.cell(row=ultima_fila+1,column=7).value = giro
                hoja_agua.cell(row=ultima_fila+1,column=6).value = direccion_completa
                hoja_agua.cell(row=ultima_fila+1,column=80).value = ciudad
                hoja_agua.cell(row=ultima_fila+1,column=47).value = ruta
                hoja_agua.cell(row=ultima_fila+1,column=8).value = rut
                
                #Segunda tabla traspasada a excel
                hoja_agua.cell(row=ultima_fila+1,column=60).value = int(consumo)
                hoja_agua.cell(row=ultima_fila+1,column=57).value = lectura_actual
                hoja_agua.cell(row=ultima_fila+1,column=56).value = int(valor_lectura_1)
                hoja_agua.cell(row=ultima_fila+1,column=58).value = lectura_anterior
                hoja_agua.cell(row=ultima_fila+1,column=65).value = int(valor_lectura_2)
                #hoja_agua.cell(row=ultima_fila+1,column=1).value = consumo_periodo
                #hoja_agua.cell(row=ultima_fila+1,column=1).value = consumo_facturado
                hoja_agua.cell(row=ultima_fila+1,column=21).value = sobreconsumo
                hoja_agua.cell(row=ultima_fila+1,column=18).value = agua_punta
                hoja_agua.cell(row=ultima_fila+1,column=63).value = int(limite)
                hoja_agua.cell(row=ultima_fila+1,column=54).value = fecha_estimada
                
                #Tercera tabla traspasada a excel
                hoja_agua.cell(row=ultima_fila+1,column=9).value = n_servicio 
                hoja_agua.cell(row=ultima_fila+1,column=11).value = vencimiento
                hoja_agua.cell(row=ultima_fila+1,column=46).value = total_a_pagar
                
                #Cuarta tabla traspasada a excel
                hoja_agua.cell(row=ultima_fila+1,column=14).value = int(cargo_fijo)
                hoja_agua.cell(row=ultima_fila+1,column=20).value = int(consumo_agua_punta)
                hoja_agua.cell(row=ultima_fila+1,column=23).value = int(sobreconsumo_agua_potable)
                #hoja_agua.cell(row=ultimaultima_fila+1_fila,column=9).value = total_servicio_mes 
                hoja_agua.cell(row=ultima_fila+1,column=81).value = int(amortizacion)
                #hoja_agua.cell(row=ultima_fila+1,column=46).value = total_venta_mes
                hoja_agua.cell(row=ultima_fila+1,column=42).value = int(valor_neto)
                hoja_agua.cell(row=ultima_fila+1,column=43).value = int(iva)
                #hoja_agua.cell(row=ultima_fila+1,column=11).value = total_docto
                hoja_agua.cell(row=ultima_fila+1,column=46).value = int(total_a_pagar_final)
                
                #Quinta tabla traspasada a excel
                hoja_agua.cell(row=ultima_fila+1,column=35).value = tarifa_publicada 
                hoja_agua.cell(row=ultima_fila+1,column=50).value = int(grupo_tarifario)
                #hoja_agua.cell(row=ultima_fila+1,column=23).value = cargo_fijo_final
                hoja_agua.cell(row=ultima_fila+1,column=15).value = m3_agua_potable
                hoja_agua.cell(row=ultima_fila+1,column=18).value = m3_agua_potable_punta
                hoja_agua.cell(row=ultima_fila+1,column=21).value = m3_agua_potable_sobreconsumo

                hoja_agua.cell(row=ultima_fila+1,column=25).value = m3_alcantarillado 
                hoja_agua.cell(row=ultima_fila+1,column=28).value = m3_tratamiento
                hoja_agua.cell(row=ultima_fila+1,column=61).value = int(factor_cobro)
                hoja_agua.cell(row=ultima_fila+1,column=52).value = int(n_medidor)
                hoja_agua.cell(row=ultima_fila+1,column=53).value = int(diametro_medidor)
                
                libro.save(output_path+'/'+'Formato Planilla.xlsx')
                    
        #         #Copiamos el archivo a la carpeta outpu con el nombre que corresponde
        #         shutil.copy(archivo, output_path+'1-'+factura_elec+'.pdf')
        #         print('-----')
    
        # #Obtenemos los archivos de la carpeta input
        # archivos_en_carpeta = os.listdir(folder_path)

        # # Iterar sobre los archivos y eliminarlos
        # for archivo in archivos_en_carpeta:
        #     ruta_archivo = os.path.join(folder_path, archivo)
        #     if os.path.isfile(ruta_archivo):
        #         os.remove(ruta_archivo)
